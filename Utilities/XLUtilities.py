import openpyxl


def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    rows=workbook[sheetName]
    return rows.max_row

def readValues(file,sheetName,rows,columns):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(rows,columns).value